import os
import re
import json
import datetime
from io import BytesIO
from collections import Counter

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from langchain_huggingface import HuggingFaceEmbeddings
from langchain_chroma import Chroma
from langchain_ollama import ChatOllama
from langchain_core.prompts import ChatPromptTemplate
from langchain_classic.chains.combine_documents import create_stuff_documents_chain

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="MobiTrace Pro — Forensic Intelligence Hub",
    layout="wide",
    page_icon="🔍",
)

st.markdown("""
<style>
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #1e1e2e;
        border-radius: 8px 8px 0 0;
        padding: 8px 20px;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] { background-color: #313244; }
    .metric-card {
        background: #1e1e2e;
        border: 1px solid #313244;
        border-radius: 10px;
        padding: 16px;
        margin-bottom: 10px;
    }
    .severity-high   { color: #f38ba8; font-weight: bold; }
    .severity-medium { color: #fab387; font-weight: bold; }
    .severity-low    { color: #a6e3a1; font-weight: bold; }
    .finding-box {
        background: #181825;
        border-left: 4px solid #89b4fa;
        border-radius: 0 8px 8px 0;
        padding: 12px 16px;
        margin: 8px 0;
        font-family: monospace;
        font-size: 13px;
    }
</style>
""", unsafe_allow_html=True)

st.title("🔍 MobiTrace Pro — Forensic Intelligence Hub")
st.caption("AI-powered Android forensic analysis · Communication · Location · Patterns · Recommendations")

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
INPUT_DIR = "./data_input"
DB_DIR    = "./chroma_db"
os.makedirs(INPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
#  MODEL LOADING (cached)
# ─────────────────────────────────────────────
@st.cache_resource
def load_heavy_models():
    with st.spinner("Loading embedding model…"):
        embeddings_model = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")
    with st.spinner("Connecting to Ollama…"):
        llm_model = ChatOllama(model="llama3", temperature=0.0)
    return embeddings_model, llm_model

embeddings, local_llm = load_heavy_models()

# ─────────────────────────────────────────────
#  VECTOR STORE
# ─────────────────────────────────────────────
if os.path.exists(DB_DIR):
    vector_store   = Chroma(persist_directory=DB_DIR, embedding_function=embeddings)
    base_retriever = vector_store.as_retriever(search_kwargs={"k": 12})
else:
    vector_store   = Chroma.from_texts(["Initial setup context"], embeddings, persist_directory=DB_DIR)
    base_retriever = vector_store.as_retriever(search_kwargs={"k": 1})

# ─────────────────────────────────────────────
#  SYSTEM PROMPTS — one per analysis module
# ─────────────────────────────────────────────

PROMPTS = {

    "communication": (
        "You are a Digital Forensic Analyst specialising in mobile communication evidence.\n"
        "Using ONLY the provided context, produce a structured Communication Analysis report.\n\n"
        "OUTPUT FORMAT — use exactly these section headers:\n"
        "## TOP CONTACTS\n"
        "List the top contacts by message frequency. Format: [ID: X] Contact | Count | Direction\n\n"
        "## COMMUNICATION TIMELINE\n"
        "Key communication events in chronological order. Cite every entry with [ID: X].\n\n"
        "## MESSAGING PATTERNS\n"
        "Note unusual hours, burst activity, one-sided conversations. Cite with [ID: X].\n\n"
        "## DELETED / GAP EVIDENCE\n"
        "Flag any missing Row IDs or gaps in sequence that may indicate deletion. Cite with [ID: X].\n\n"
        "RULES: Every finding must cite [ID: X]. If no evidence exists for a section, write 'No evidence found.'\n\n"
        "Context:\n{context}"
    ),

    "location": (
        "You are a Digital Forensic Analyst specialising in location and geospatial evidence.\n"
        "Using ONLY the provided context, produce a structured Location Findings report.\n\n"
        "OUTPUT FORMAT — use exactly these section headers:\n"
        "## GPS COORDINATES FOUND\n"
        "List all latitude/longitude pairs with source. Cite with [ID: X].\n\n"
        "## LOCATION MENTIONS IN MESSAGES\n"
        "Extract any place names, addresses, or meeting points mentioned. Cite with [ID: X].\n\n"
        "## MOVEMENT TIMELINE\n"
        "Reconstruct suspect movement in chronological order from location evidence. Cite with [ID: X].\n\n"
        "## SUSPICIOUS LOCATION ACTIVITY\n"
        "Flag unusual locations (late-night meeting spots, known high-risk areas, repeated covert locations). "
        "Cite with [ID: X].\n\n"
        "RULES: Every finding must cite [ID: X]. If no evidence exists for a section, write 'No evidence found.'\n\n"
        "Context:\n{context}"
    ),

    "patterns": (
        "You are a Digital Forensic Analyst specialising in behavioural anomaly detection.\n"
        "Using ONLY the provided context, produce a structured Suspicious Patterns report.\n\n"
        "OUTPUT FORMAT — use exactly these section headers:\n"
        "## CODED LANGUAGE DETECTION\n"
        "Identify phrases that appear to use coded, ambiguous, or evasive language. "
        "Assign a suspicion level (HIGH / MEDIUM / LOW). Cite with [ID: X].\n\n"
        "## TEMPORAL ANOMALIES\n"
        "Flag messages or calls at unusual hours (10 PM – 5 AM), rapid bursts, or sudden silence periods. "
        "Cite with [ID: X].\n\n"
        "## NETWORK ANOMALIES\n"
        "Identify new/unknown contacts appearing suddenly, unidirectional communication, "
        "or contacts referenced but not in contact book. Cite with [ID: X].\n\n"
        "## DELETION / ANTI-FORENSIC INDICATORS\n"
        "Flag Row ID gaps, suspiciously clean message threads, or sudden app uninstalls. "
        "Cite with [ID: X].\n\n"
        "## SEVERITY MATRIX\n"
        "Assign an overall Forensic Risk Index score (1–10) with a one-line justification.\n\n"
        "RULES: Every finding must cite [ID: X]. If no evidence exists for a section, write 'No evidence found.'\n\n"
        "Context:\n{context}"
    ),

    "recommendations": (
        "You are a Senior Digital Forensic Investigator writing actionable recommendations for a law enforcement team.\n"
        "Based ONLY on the provided context evidence, produce a structured Recommendations report.\n\n"
        "OUTPUT FORMAT — use exactly these section headers:\n"
        "## IMMEDIATE ACTIONS (Next 24 Hours)\n"
        "List urgent investigative steps the team should take right now. Number each action.\n\n"
        "## ADDITIONAL WARRANT TARGETS\n"
        "List specific cloud accounts, devices, or services that should be subpoenaed based on the evidence. "
        "Cite the evidence that justifies each warrant with [ID: X].\n\n"
        "## EVIDENCE PRESERVATION PRIORITIES\n"
        "List which evidence items are most at risk of loss and must be preserved immediately. "
        "Cite with [ID: X].\n\n"
        "## INVESTIGATIVE LEADS\n"
        "List unresolved questions and suggested follow-up lines of enquiry.\n\n"
        "## COURT READINESS ASSESSMENT\n"
        "State which findings are court-admissible based on citation traceability, "
        "and flag any findings that need additional corroboration.\n\n"
        "RULES: Every finding that references evidence must cite [ID: X]. Be precise and actionable.\n\n"
        "Context:\n{context}"
    ),

    "chat": (
        "You are a strict digital forensic data auditor inspecting real-time messaging logs.\n"
        "CRITICAL CHRONOLOGY RULE:\n"
        "1. Context fragments with low row markers (row_id 1, row 2) represent the MOST RECENT information.\n"
        "2. Focus on the most recent target event. Ignore older trailing rows (row 2450+) once found.\n"
        "3. Every claim you make MUST be cited with [ID: X]. If you cannot find evidence, say so.\n\n"
        "Context:\n{context}"
    ),
}

# ─────────────────────────────────────────────
#  CORE RAG ENGINE
# ─────────────────────────────────────────────
def extract_row_number(doc):
    content = doc.page_content.lower()
    for i in range(1, 200):
        if f"row {i}" in content or f"line {i}" in content or f"row_id {i}" in content:
            return i
    return 9999

def build_chain(prompt_key: str):
    sys_prompt = PROMPTS[prompt_key]
    prompt_template = ChatPromptTemplate.from_messages([
        ("system", sys_prompt),
        ("human", "{input}"),
    ])
    return create_stuff_documents_chain(local_llm, prompt_template)

def rag_invoke(query_text: str, prompt_key: str, k: int = 12) -> str:
    retriever = vector_store.as_retriever(search_kwargs={"k": k})
    docs      = retriever.invoke(query_text)

    # Chronological sort
    sorted_docs = sorted(docs, key=extract_row_number)

    # Early termination — drop ancient rows when fresh ones are found
    optimized_docs = []
    for doc in sorted_docs:
        optimized_docs.append(doc)
        if extract_row_number(doc) < 50:
            break

    chain    = build_chain(prompt_key)
    response = chain.invoke({"input": query_text, "context": optimized_docs})
    return response

# ─────────────────────────────────────────────
#  HALLUCINATION FIREWALL
# ─────────────────────────────────────────────
def validate_citations(analysis_text: str, docs: list) -> tuple[str, list]:
    """
    Cross-checks every [ID: X] in AI output against retrieved doc row IDs.
    Returns (validated_text, list_of_phantom_ids).
    """
    all_content = " ".join([d.page_content.lower() for d in docs])
    cited_ids   = [int(x) for x in re.findall(r'\[ID:\s*(\d+)\]', analysis_text)]
    phantoms    = []

    for cid in cited_ids:
        if f"row {cid}" not in all_content and \
           f"row_id {cid}" not in all_content and \
           f"line {cid}" not in all_content:
            phantoms.append(cid)
            analysis_text = analysis_text.replace(
                f"[ID: {cid}]",
                f"[ID: {cid} ⚠️ VERIFY]"
            )
    return analysis_text, phantoms

# ─────────────────────────────────────────────
#  PDF GENERATION — enhanced with sections
# ─────────────────────────────────────────────
def generate_pdf(sections: dict, case_id: str = "MobiTrace Report") -> BytesIO:
    buffer = BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=50, leftMargin=50,
        topMargin=50,   bottomMargin=50
    )
    styles  = getSampleStyleSheet()
    story   = []

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'],
        fontSize=20, textColor=colors.HexColor('#1a1a2e'),
        spaceAfter=6
    )
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'],
        fontSize=9,  textColor=colors.grey, spaceAfter=16
    )
    h2_style = ParagraphStyle(
        'H2', parent=styles['Heading2'],
        fontSize=14, textColor=colors.HexColor('#16213e'),
        spaceBefore=14, spaceAfter=6
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['BodyText'],
        fontSize=10, leading=15, spaceAfter=6
    )
    citation_style = ParagraphStyle(
        'Citation', parent=styles['Code'],
        fontSize=9, backColor=colors.HexColor('#f0f4ff'),
        borderPadding=4, spaceAfter=4, leading=13
    )

    # Header
    story.append(Paragraph(f"🔍 {case_id}", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} · MobiTrace Pro",
        meta_style
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cccccc')))
    story.append(Spacer(1, 12))

    SECTION_LABELS = {
        "communication":   "📞 Communication Analysis",
        "location":        "📍 Location Findings",
        "patterns":        "⚠️ Suspicious Patterns",
        "recommendations": "📋 Investigative Recommendations",
    }

    for key, content in sections.items():
        if not content or not content.strip():
            continue

        story.append(Paragraph(SECTION_LABELS.get(key, key.title()), h2_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dddddd')))
        story.append(Spacer(1, 6))

        for line in content.split('\n'):
            line = line.strip()
            if not line:
                story.append(Spacer(1, 4))
                continue
            if line.startswith('##'):
                story.append(Paragraph(
                    line.replace('##', '').strip(),
                    ParagraphStyle('Sub', parent=styles['Heading3'],
                                   fontSize=11, textColor=colors.HexColor('#0d3b66'),
                                   spaceBefore=10, spaceAfter=4)
                ))
            elif '[ID:' in line:
                story.append(Paragraph(line.replace('<', '&lt;').replace('>', '&gt;'), citation_style))
            else:
                story.append(Paragraph(line.replace('**', '').replace('*', ''), body_style))

        story.append(Spacer(1, 16))

    # Officer sign-off table
    story.append(Paragraph("✍️ Investigator Verification Sign-Off", h2_style))
    sign_data = [
        ["Finding #", "Summary", "Verdict", "Officer Initials"],
        *[["", "", "☐ Accept  ☐ Reject", ""] for _ in range(6)]
    ]
    sign_table = Table(sign_data, colWidths=[0.6*inch, 3.4*inch, 1.5*inch, 1.2*inch])
    sign_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('MINROWHEIGHT', (0, 1), (-1, -1), 24),
    ]))
    story.append(sign_table)

    doc.build(story)
    buffer.seek(0)
    return buffer

# ─────────────────────────────────────────────
#  EXCEL GENERATION — enhanced
# ─────────────────────────────────────────────
def generate_excel(sections: dict) -> BytesIO:
    buffer = BytesIO()
    wb     = Workbook()

    SHEET_LABELS = {
        "communication":   "Communication Analysis",
        "location":        "Location Findings",
        "patterns":        "Suspicious Patterns",
        "recommendations": "Recommendations",
    }

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    sub_fill    = PatternFill("solid", fgColor="e8eaf6")
    cite_fill   = PatternFill("solid", fgColor="fff3e0")

    first = True
    for key, content in sections.items():
        if not content or not content.strip():
            continue
        label = SHEET_LABELS.get(key, key.title())
        ws = wb.active if first else wb.create_sheet()
        ws.title = label[:31]
        first = False

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 20

        # Sheet header
        ws['A1'] = "MobiTrace Pro — " + label
        ws['A1'].font      = Font(bold=True, size=13, color="1a1a2e")
        ws['A1'].alignment = Alignment(horizontal='left')
        ws.merge_cells('A1:C1')

        ws['A2'] = "Generated"
        ws['B2'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Column headers
        for col, hdr in enumerate(["Section / Finding", "Content", "Citation IDs"], start=1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row_num = 5
        current_section = ""
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
            # Extract citation IDs from line
            cids = ", ".join(re.findall(r'\[ID:\s*(\d+)\]', line)) or ""

            if line.startswith('##'):
                current_section = line.replace('##', '').strip()
                cell_a = ws.cell(row=row_num, column=1, value=current_section)
                cell_b = ws.cell(row=row_num, column=2, value="")
                cell_c = ws.cell(row=row_num, column=3, value="")
                for c in [cell_a, cell_b, cell_c]:
                    c.fill = sub_fill
                    c.font = Font(bold=True, size=10)
            else:
                clean_line = re.sub(r'\[ID:\s*\d+\]', '', line).strip()
                cell_a = ws.cell(row=row_num, column=1, value=current_section)
                cell_b = ws.cell(row=row_num, column=2, value=clean_line)
                cell_c = ws.cell(row=row_num, column=3, value=cids)
                if cids:
                    for c in [cell_a, cell_b, cell_c]:
                        c.fill = cite_fill
                cell_b.alignment = Alignment(wrap_text=True)

            row_num += 1

    # Remove default empty sheet if we added named ones
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']

    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─────────────────────────────────────────────
#  SESSION STATE INIT
# ─────────────────────────────────────────────
for key in ["messages", "comm_result", "loc_result", "patterns_result", "reco_result"]:
    if key not in st.session_state:
        st.session_state[key] = [] if key == "messages" else ""

# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Case Settings")
    case_id = st.text_input("Case ID", value="CID-2026-001")
    investigator = st.text_input("Investigator", value="")
    st.divider()

    st.header("📋 Export Full Report")
    all_sections = {
        "communication":   st.session_state.comm_result,
        "location":        st.session_state.loc_result,
        "patterns":        st.session_state.patterns_result,
        "recommendations": st.session_state.reco_result,
    }
    has_content = any(v.strip() for v in all_sections.values())

    if has_content:
        report_title = f"MobiTrace Forensic Report — {case_id}"
        st.download_button(
            "📥 Download Full PDF Report",
            data=generate_pdf(all_sections, report_title),
            file_name=f"{case_id}_forensic_report.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
        st.download_button(
            "📥 Download Excel Evidence Matrix",
            data=generate_excel(all_sections),
            file_name=f"{case_id}_evidence_matrix.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("Run at least one analysis module to enable export.")

    st.divider()
    st.header("🔥 Hallucination Guard")
    st.caption("Every [ID: X] citation is cross-checked against retrieved evidence before display.")

# ─────────────────────────────────────────────
#  MAIN TABS
# ─────────────────────────────────────────────
tab_chat, tab_comm, tab_loc, tab_patterns, tab_reco = st.tabs([
    "💬 Evidence Query",
    "📞 Communication Analysis",
    "📍 Location Findings",
    "⚠️ Suspicious Patterns",
    "📋 Recommendations",
])

# ══════════════════════════════════════════════
#  TAB 1 — EVIDENCE QUERY CHAT
# ══════════════════════════════════════════════
with tab_chat:
    st.subheader("💬 Natural Language Evidence Query")
    st.caption("Ask anything about the extracted evidence. Every answer is citation-anchored.")

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    if user_query := st.chat_input("Ask about your evidence… e.g. 'Who did the suspect call most after midnight?'"):
        with st.chat_message("user"):
            st.markdown(user_query)
        st.session_state.messages.append({"role": "user", "content": user_query})

        with st.chat_message("assistant"):
            with st.spinner("Scanning evidence with chronological priority…"):
                try:
                    answer = rag_invoke(user_query, "chat", k=12)
                    # Validate citations
                    docs   = vector_store.as_retriever(search_kwargs={"k": 12}).invoke(user_query)
                    answer, phantoms = validate_citations(answer, docs)
                    if phantoms:
                        answer += f"\n\n> ⚠️ **Hallucination Guard:** {len(phantoms)} citation(s) could not be verified in evidence: {phantoms}"
                except Exception as e:
                    answer = f"❌ Error: {e}"
                st.markdown(answer)
        st.session_state.messages.append({"role": "assistant", "content": answer})

    if st.session_state.messages:
        chat_log = "\n\n".join([
            f"**{m['role'].upper()}:** {m['content']}"
            for m in st.session_state.messages
        ])
        st.download_button(
            "📥 Export Chat Log as PDF",
            data=generate_pdf({"chat": chat_log}, f"Query Log — {case_id}"),
            file_name=f"{case_id}_query_log.pdf",
            mime="application/pdf",
        )

# ══════════════════════════════════════════════
#  TAB 2 — COMMUNICATION ANALYSIS
# ══════════════════════════════════════════════
with tab_comm:
    st.subheader("📞 Communication Analysis")
    st.caption("Extracts contact frequency, message timeline, directional patterns, and deletion gaps.")

    col_a, col_b = st.columns([3, 1])
    with col_a:
        comm_query = st.text_area(
            "Focus query (optional — leave blank for full analysis)",
            value="Analyse all SMS and call log communications. Identify top contacts, "
                  "unusual messaging times, one-sided conversations, and any gaps in Row IDs "
                  "that may indicate deletion.",
            height=90,
            key="comm_query_input"
        )
    with col_b:
        st.markdown("<br>", unsafe_allow_html=True)
        run_comm = st.button("▶ Run Analysis", key="run_comm", use_container_width=True, type="primary")

    if run_comm:
        with st.spinner("Running communication analysis…"):
            try:
                result = rag_invoke(comm_query, "communication", k=15)
                docs   = vector_store.as_retriever(search_kwargs={"k": 15}).invoke(comm_query)
                result, phantoms = validate_citations(result, docs)
                if phantoms:
                    result += f"\n\n> ⚠️ **Hallucination Guard:** Citations {phantoms} could not be verified."
                st.session_state.comm_result = result
            except Exception as e:
                st.session_state.comm_result = f"❌ Error: {e}"

    if st.session_state.comm_result:
        st.markdown("---")
        # Parse and display sections with styled boxes
        sections_text = st.session_state.comm_result.split("##")
        for section in sections_text:
            if not section.strip():
                continue
            lines = section.strip().split('\n')
            title = lines[0].strip()
            body  = '\n'.join(lines[1:]).strip()
            with st.expander(f"📌 {title}", expanded=True):
                # Highlight citation lines
                for line in body.split('\n'):
                    if '[ID:' in line:
                        st.markdown(f'<div class="finding-box">{line}</div>', unsafe_allow_html=True)
                    elif line.strip():
                        st.markdown(line)

        st.download_button(
            "📥 Export Communication Report",
            data=generate_pdf({"communication": st.session_state.comm_result},
                              f"Communication Analysis — {case_id}"),
            file_name=f"{case_id}_communication.pdf",
            mime="application/pdf",
        )

# ══════════════════════════════════════════════
#  TAB 3 — LOCATION FINDINGS
# ══════════════════════════════════════════════
with tab_loc:
    st.subheader("📍 Location Findings")
    st.caption("Extracts GPS coordinates, location mentions in messages, and movement timelines.")

    col_a, col_b = st.columns([3, 1])
    with col_a:
        loc_query = st.text_area(
            "Focus query (optional)",
            value="Extract all GPS coordinates, place names, addresses, and location references "
                  "from the evidence. Build a movement timeline and flag any suspicious locations "
                  "such as late-night meeting points or repeated covert spots.",
            height=90,
            key="loc_query_input"
        )
    with col_b:
        st.markdown("<br>", unsafe_allow_html=True)
        run_loc = st.button("▶ Run Analysis", key="run_loc", use_container_width=True, type="primary")

    if run_loc:
        with st.spinner("Extracting location evidence…"):
            try:
                result = rag_invoke(loc_query, "location", k=15)
                docs   = vector_store.as_retriever(search_kwargs={"k": 15}).invoke(loc_query)
                result, phantoms = validate_citations(result, docs)
                if phantoms:
                    result += f"\n\n> ⚠️ **Hallucination Guard:** Citations {phantoms} could not be verified."
                st.session_state.loc_result = result
            except Exception as e:
                st.session_state.loc_result = f"❌ Error: {e}"

    if st.session_state.loc_result:
        st.markdown("---")
        # Check if any GPS coordinates were found — display a map hint
        coords = re.findall(
            r'(-?\d{1,3}\.\d+)[,\s]+(-?\d{1,3}\.\d+)',
            st.session_state.loc_result
        )
        if coords:
            st.info(f"📌 {len(coords)} GPS coordinate(s) detected. "
                    "Export the report to see them in context.")

        sections_text = st.session_state.loc_result.split("##")
        for section in sections_text:
            if not section.strip():
                continue
            lines = section.strip().split('\n')
            title = lines[0].strip()
            body  = '\n'.join(lines[1:]).strip()
            with st.expander(f"📍 {title}", expanded=True):
                for line in body.split('\n'):
                    if '[ID:' in line:
                        st.markdown(f'<div class="finding-box">{line}</div>', unsafe_allow_html=True)
                    elif line.strip():
                        st.markdown(line)

        st.download_button(
            "📥 Export Location Report",
            data=generate_pdf({"location": st.session_state.loc_result},
                              f"Location Findings — {case_id}"),
            file_name=f"{case_id}_location.pdf",
            mime="application/pdf",
        )

# ══════════════════════════════════════════════
#  TAB 4 — SUSPICIOUS PATTERNS
# ══════════════════════════════════════════════
with tab_patterns:
    st.subheader("⚠️ Suspicious Patterns Detection")
    st.caption("Detects coded language, temporal anomalies, network anomalies, and anti-forensic indicators.")

    col_a, col_b = st.columns([3, 1])
    with col_a:
        pat_query = st.text_area(
            "Focus query (optional)",
            value="Detect all suspicious patterns including coded or evasive language, "
                  "unusual timing (late night or burst messaging), unexpected contacts, "
                  "and any signs of deliberate evidence deletion such as Row ID gaps. "
                  "Assign a severity level to each finding and calculate a Forensic Risk Index.",
            height=110,
            key="pat_query_input"
        )
    with col_b:
        st.markdown("<br>", unsafe_allow_html=True)
        run_pat = st.button("▶ Run Analysis", key="run_pat", use_container_width=True, type="primary")

    if run_pat:
        with st.spinner("Scanning for suspicious patterns…"):
            try:
                result = rag_invoke(pat_query, "patterns", k=15)
                docs   = vector_store.as_retriever(search_kwargs={"k": 15}).invoke(pat_query)
                result, phantoms = validate_citations(result, docs)
                if phantoms:
                    result += f"\n\n> ⚠️ **Hallucination Guard:** Citations {phantoms} could not be verified."
                st.session_state.patterns_result = result
            except Exception as e:
                st.session_state.patterns_result = f"❌ Error: {e}"

    if st.session_state.patterns_result:
        st.markdown("---")

        # Extract and display Forensic Risk Index as a metric
        risk_match = re.search(
            r'forensic risk index[^\d]*(\d+)',
            st.session_state.patterns_result, re.IGNORECASE
        )
        if risk_match:
            score = int(risk_match.group(1))
            colour = "🔴" if score >= 7 else "🟠" if score >= 4 else "🟢"
            st.metric(
                label="Forensic Risk Index",
                value=f"{colour} {score} / 10",
                delta="High Priority" if score >= 7 else
                      "Medium Priority" if score >= 4 else "Low Priority"
            )

        # Count severity labels
        high_count   = st.session_state.patterns_result.upper().count("HIGH")
        medium_count = st.session_state.patterns_result.upper().count("MEDIUM")
        low_count    = st.session_state.patterns_result.upper().count("LOW")
        if high_count or medium_count or low_count:
            c1, c2, c3 = st.columns(3)
            c1.metric("🔴 HIGH findings",   high_count)
            c2.metric("🟠 MEDIUM findings", medium_count)
            c3.metric("🟢 LOW findings",    low_count)

        sections_text = st.session_state.patterns_result.split("##")
        for section in sections_text:
            if not section.strip():
                continue
            lines = section.strip().split('\n')
            title = lines[0].strip()
            body  = '\n'.join(lines[1:]).strip()
            with st.expander(f"⚠️ {title}", expanded=True):
                for line in body.split('\n'):
                    if '[ID:' in line:
                        st.markdown(f'<div class="finding-box">{line}</div>', unsafe_allow_html=True)
                    elif 'HIGH' in line.upper():
                        st.markdown(f'<span class="severity-high">🔴 {line}</span>', unsafe_allow_html=True)
                    elif 'MEDIUM' in line.upper():
                        st.markdown(f'<span class="severity-medium">🟠 {line}</span>', unsafe_allow_html=True)
                    elif 'LOW' in line.upper():
                        st.markdown(f'<span class="severity-low">🟢 {line}</span>', unsafe_allow_html=True)
                    elif line.strip():
                        st.markdown(line)

        st.download_button(
            "📥 Export Patterns Report",
            data=generate_pdf({"patterns": st.session_state.patterns_result},
                              f"Suspicious Patterns — {case_id}"),
            file_name=f"{case_id}_patterns.pdf",
            mime="application/pdf",
        )

# ══════════════════════════════════════════════
#  TAB 5 — RECOMMENDATIONS
# ══════════════════════════════════════════════
with tab_reco:
    st.subheader("📋 Investigative Recommendations")
    st.caption("AI-generated actionable steps, warrant targets, evidence priorities, and court readiness assessment.")

    col_a, col_b = st.columns([3, 1])
    with col_a:
        reco_query = st.text_area(
            "Focus query (optional)",
            value="Based on all the extracted evidence, generate a complete set of investigative "
                  "recommendations including immediate actions, additional warrant targets, "
                  "evidence preservation priorities, investigative leads, and a court readiness "
                  "assessment for each key finding.",
            height=110,
            key="reco_query_input"
        )
    with col_b:
        st.markdown("<br>", unsafe_allow_html=True)
        run_reco = st.button("▶ Run Analysis", key="run_reco", use_container_width=True, type="primary")

    if run_reco:
        with st.spinner("Generating investigative recommendations…"):
            try:
                result = rag_invoke(reco_query, "recommendations", k=20)
                docs   = vector_store.as_retriever(search_kwargs={"k": 20}).invoke(reco_query)
                result, phantoms = validate_citations(result, docs)
                if phantoms:
                    result += f"\n\n> ⚠️ **Hallucination Guard:** Citations {phantoms} could not be verified."
                st.session_state.reco_result = result
            except Exception as e:
                st.session_state.reco_result = f"❌ Error: {e}"

    if st.session_state.reco_result:
        st.markdown("---")

        sections_text = st.session_state.reco_result.split("##")
        SECTION_ICONS = {
            "IMMEDIATE":    "🚨",
            "WARRANT":      "📜",
            "PRESERVATION": "🛡️",
            "LEADS":        "🔎",
            "COURT":        "⚖️",
        }
        for section in sections_text:
            if not section.strip():
                continue
            lines  = section.strip().split('\n')
            title  = lines[0].strip()
            body   = '\n'.join(lines[1:]).strip()
            icon   = next((v for k, v in SECTION_ICONS.items() if k in title.upper()), "📌")
            with st.expander(f"{icon} {title}", expanded=True):
                for line in body.split('\n'):
                    if '[ID:' in line:
                        st.markdown(f'<div class="finding-box">{line}</div>', unsafe_allow_html=True)
                    elif line.strip().startswith(tuple('0123456789')):
                        st.markdown(f"**{line}**")
                    elif line.strip():
                        st.markdown(line)

        st.download_button(
            "📥 Export Recommendations Report",
            data=generate_pdf({"recommendations": st.session_state.reco_result},
                              f"Recommendations — {case_id}"),
            file_name=f"{case_id}_recommendations.pdf",
            mime="application/pdf",
        )
